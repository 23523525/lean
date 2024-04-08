import asyncio

import openpyxl

from DataBase.dbModels import FiveHundredWords, ThreeThousenWords

# перенос данных их excel в базу данных
async def from_excel_to_db_500():
    wb = openpyxl.load_workbook('500words.xlsx')
    sheet = wb.active

    # добавляем каждую строку excel в базу данных
    for row in sheet.iter_rows():
        if row[0].value != None:
            await FiveHundredWords.add_word(
                english=str(row[0].value) + ' ' + str(row[1].value),
                russian=str(row[2].value)
            )

    wb_2 = openpyxl.load_workbook('3000words.xlsx')
    sheet_2 = wb_2.active

    for row in sheet_2.iter_rows():
        if row[0].value != None:
            await ThreeThousenWords.add_word(
                english=str(row[0].value) + ' ' + str(row[1].value),
                russian=str(row[2].value)
            )

asyncio.run(from_excel_to_db_500())